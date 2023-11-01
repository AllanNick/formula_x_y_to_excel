from openpyxl import load_workbook

alphas = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

def write_matrix(b_name, obj, output_name):
    wb = load_workbook(b_name)

    ws = wb.active

    line_index = 1
    #colm_index = get_colm(colm)
    l = 0
    c = 0
    for line_ele in obj:
        l += 1
        for colm_ele in line_ele:
            c += 1
            ws.cell(row=l, column=c, value=colm_ele)
        c = 0

    wb.save(output_name)
